import streamlit as st
import pandas as pd
import random
import io
import openpyxl
from openpyxl.utils import get_column_letter

# --- Page Configuration ---
st.set_page_config(
    page_title="OBE Intelligence Pro",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- Custom Styling ---
st.markdown("""
    <style>
    .main { background-color: #f8f9fa; }
    .stButton>button { width: 100%; border-radius: 5px; height: 3em; background-color: #007bff; color: white; font-weight: bold; }
    .stDownloadButton>button { width: 100%; border-radius: 5px; background-color: #28a745; color: white; font-weight: bold; }
    div[data-testid="stExpander"] { border: 1px solid #e1e4e8; border-radius: 8px; background-color: white; }
    h1 { color: #1e3a8a; font-family: 'Segoe UI', sans-serif; }
    .stTabs [data-baseweb="tab-list"] { gap: 24px; }
    .stTabs [data-baseweb="tab"] { height: 50px; white-space: pre-wrap; background-color: #f0f2f6; border-radius: 5px 5px 0 0; padding: 10px 20px; }
    .stTabs [aria-selected="true"] { background-color: #007bff !important; color: white !important; }
    </style>
    """, unsafe_allow_html=True)

# --- Core Logic ---
def distribute_marks(obtained, clo_max_marks):
    """Robust mark distribution logic with error handling."""
    try:
        obtained = float(obtained)
        if pd.isna(obtained) or obtained <= 0:
            return [0.0] * len(clo_max_marks)
        
        # Check if student marks exceed total possible
        total_possible = sum(clo_max_marks)
        if obtained > total_possible:
            obtained = total_possible
            
        allocated = [0.0] * len(clo_max_marks)
        remaining = obtained
        
        # Safety counter to prevent infinite loops
        iterations = 0
        while remaining > 0.001 and iterations < 5000:
            available_indices = [i for i, max_val in enumerate(clo_max_marks) if allocated[i] < max_val]
            if not available_indices: break
                
            idx = random.choice(available_indices)
            space_left = clo_max_marks[idx] - allocated[idx]
            
            # Use random increments (0.5 or 1.0)
            step = random.choice([0.5, 1.0])
            chunk = min(step, remaining, space_left)
            
            allocated[idx] += chunk
            remaining -= chunk
            iterations += 1
            
        return [round(m, 2) for m in allocated]
    except Exception:
        return [0.0] * len(clo_max_marks)

def main():
    st.title("🎯 OBE Marks Intelligence Pro")
    st.markdown("##### Transform portal exports into structured CLO data and map them to official templates.")

    # Sidebar Instructions
    with st.sidebar:
        st.header("Help & Instructions")
        st.info("""
        1. **Upload**: Support .csv, .xlsx, .xls (including HTML exports).
        2. **Configure**: Define your CLO max marks.
        3. **Process**: Randomly distribute marks per student.
        4. **Map**: Upload your university template to auto-fill.
        """)
        st.divider()
        st.caption("v2.0 - Universal Format Support")

    tab1, tab2 = st.tabs(["📊 1. Distribution Engine", "📄 2. Template Mapper"])

    # --- TAB 1: DISTRIBUTION ---
    with tab1:
        st.header("Step 1: Data Intake")
        uploaded_file = st.file_uploader(
            "Upload Roster / Result File", 
            type=['csv', 'xlsx', 'xls'],
            help="Upload the file from your portal. Even if it says .xls, this tool will fix format mismatches."
        )

        if uploaded_file:
            try:
                file_ext = uploaded_file.name.split('.')[-1].lower()
                
                # --- UNIVERSAL LOADER (Fixes BOF/HTML error) ---
                if file_ext == 'csv':
                    df = pd.read_csv(uploaded_file)
                else:
                    try:
                        engine = 'xlrd' if file_ext == 'xls' else 'openpyxl'
                        df = pd.read_excel(uploaded_file, engine=engine)
                    except Exception as e:
                        if "BOF" in str(e) or "html" in str(e).lower():
                            st.info("🔄 Technical format mismatch detected. Activating HTML Recovery Mode...")
                            html_tables = pd.read_html(uploaded_file)
                            df = html_tables[0]
                        else:
                            raise e

                # Basic Cleanup
                df = df.dropna(how='all').dropna(axis=1, how='all')
                
                st.success(f"✅ Successfully loaded {len(df)} records.")
                
                with st.expander("🔍 Preview Raw Portal Data"):
                    st.dataframe(df.head(10), use_container_width=True)

                st.divider()
                st.header("Step 2: Column Mapping & CLO Setup")
                cols = df.columns.tolist()
                
                c1, c2, c3 = st.columns(3)
                name_col = c1.selectbox("Name Column", cols, help="Column containing student names.")
                roll_col = c2.selectbox("Roll No Column", cols, help="Column containing Registration/IDs.")
                marks_col = c3.selectbox("Obtained Marks", cols, help="The total score column you want to split.")

                num_clos = st.number_input("Total CLOs in this Exam", min_value=1, value=3, help="Define how many CLOs were tested.")
                
                clo_max_marks = []
                clo_ui_cols = st.columns(num_clos)
                for i in range(num_clos):
                    m = clo_ui_cols[i].number_input(f"CLO {i+1} Max", min_value=0.1, value=10.0, key=f"m_{i}", help=f"Set max possible marks for CLO {i+1}")
                    clo_max_marks.append(m)

                if st.button("🚀 Generate CLO Distribution"):
                    with st.spinner("Calculating distributions..."):
                        if not pd.api.types.is_numeric_dtype(df[marks_col]):
                            # Attempt to force numeric
                            df[marks_col] = pd.to_numeric(df[marks_col], errors='coerce').fillna(0)
                        
                        results = df[marks_col].apply(lambda x: distribute_marks(x, clo_max_marks))
                        for i in range(num_clos):
                            df[f"CLO_{i+1}_GEN"] = [res[i] for res in results]
                        
                        st.session_state['processed_df'] = df
                        st.session_state['num_clos'] = num_clos
                        st.balloons()
                        st.success("Marks Distributed! Proceed to 'Template Mapper' tab.")
                        st.dataframe(df.head(), use_container_width=True)

            except Exception as e:
                st.error(f"❌ Error reading file: {e}")

    # --- TAB 2: MAPPING ---
    with tab2:
        if 'processed_df' not in st.session_state:
            st.warning("⚠️ Please process your data in the 'Distribution Engine' tab first.")
        else:
            st.header("Step 3: Official Template Integration")
            template_file = st.file_uploader(
                "Upload Your University Excel Template", 
                type=['xlsx'],
                help="Upload the blank Excel sheet where you need the marks filled. MUST be .xlsx."
            )
            
            if template_file:
                try:
                    wb_temp = openpyxl.load_workbook(template_file)
                    sheet_name = st.selectbox("Select Target Sheet", wb_temp.sheetnames)
                    sheet = wb_temp[sheet_name]
                    
                    st.info("📍 Map Excel Coordinates")
                    m1, m2, m3 = st.columns(3)
                    start_row = m1.number_input("Starting Row", min_value=1, value=5, help="Row number where the first student's name appears.")
                    name_target = m2.text_input("Name Col (Letter)", "B").upper()
                    roll_target = m3.text_input("Roll Col (Letter)", "A").upper()

                    st.write("**Map CLOs to Template Columns**")
                    clo_map_cols = st.columns(st.session_state['num_clos'])
                    clo_target_letters = []
                    for i in range(st.session_state['num_clos']):
                        let = clo_map_cols[i].text_input(f"CLO {i+1} Col", value=get_column_letter(3+i), key=f"t_{i}").upper()
                        clo_target_letters.append(let)

                    if st.button("🪄 Finalize & Map Template"):
                        with st.spinner("Writing to template..."):
                            final_df = st.session_state['processed_df']
                            for idx, row in final_df.iterrows():
                                curr = int(start_row + idx)
                                sheet[f"{name_target}{curr}"] = row[name_col]
                                sheet[f"{roll_target}{curr}"] = row[roll_col]
                                for i in range(st.session_state['num_clos']):
                                    sheet[f"{clo_target_letters[i]}{curr}"] = row[f"CLO_{i+1}_GEN"]

                            out_ptr = io.BytesIO()
                            wb_temp.save(out_ptr)
                            out_ptr.seek(0)
                            
                            st.download_button(
                                label="💾 Download Mapped OBE Result",
                                data=out_ptr,
                                file_name="OBE_Final_Mapping.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                except Exception as e:
                    st.error(f"❌ Template Error: {e}")

if __name__ == "__main__":
    main()
